# Program for modifying excel tables of categories and products for Prestashop import

# import setup_ps_prep


# Definitions of all functions


import os
import openpyxl

print('imported openpyxl, os and setup.py file')

# newTable = openpyxl.load_workbook(input()) [only when multiple tables in folder]
cats_wb = openpyxl.load_workbook('files\cats.xlsx')
cats = cats_wb['Sheet1']



# Beginning Categories Functions

# Rename all cell elements in Column 'H' to have ' von Nordent' behind them+++++++
def cat_meta_title():
    
    for x in range(1, len(cats['H'])):
        oldTitle = str(cats['H'][x].value)
        newTitle = str(oldTitle) + ' von Nordent'
        cats['H'][x].value = newTitle
    print('1.) Column H: Meta Titles adjusted.')

# Rename parent_id according to categories_name column 'G' of respective parent++++++
def cat_corr_parents():
    
    for x in range(1, len(cats['C'])):
        oldParent_id = int(cats['C'][x].value)
        if oldParent_id == 0:
            cats['C'][x].value = str('Home')
            # print(str(cats['C'][x].value))
        else:
            cats['C'][x].value = str(cats['G'][oldParent_id - 1].value)
            # print(str(cats['C'][x].value))
    print('2.) Column C: Category Parents changed from ID# to actual parent names.')

# Rename categories_meta_keywords to friendly URL title ++++++++
def cat_friendly_urls():
    
    for x in range(1, len(cats['J'])):
        if "," in str(cats['J'][x].value):
            cats['J'][x].value = str(cats['J'][x].value).replace(",", "")
        if " / " in str(cats['J'][x].value):
            cats['J'][x].value = str(cats['J'][x].value).replace(" / ", "")
        if " - " in str(cats['J'][x].value):
            cats['J'][x].value = str(cats['J'][x].value).replace(" - ", " ")
        cats['J'][x].value = str('Nordent ' + cats['J'][x].value)
        # print(str(cats['J'][x].value))
    print('3.) Column J: can now be used as friendly URL links.')

# Rename img links to have rel. path (/html/ukens-dental/img/nordent_de_cat_images/+)
def cat_img_links():
    
    for x in range(1, len(cats['F'])):
        oldImg = str(cats['F'][x].value)
        cats['F'][x].value = str('https://ukens-dental.de/img/nordent_de_cat_images/' + oldImg)
        # print(str(cats['F'][x].value) + '.....' + str(cats['A'][x].value))
    print('4.) Column F: img links point to: https://ukens-dental.de/img/nordent_de_cat_images/ + xxx.jpg.')

# Rewrite Categories to be 1xxx, so Nordent is in thousands, whereas Calset is then 2000
def cat_corr_cat_ids():
    
    for x in range(1, len(cats['A'])):
        cats['A'][x].value = int(cats['A'][x].value) + 1000
        # print(str(cats['A'][x].value))
    print('5.) Column A: Category IDs have been +1000, now range from 1002 to 1126.')


# Make Meta Descriptions same as Category Titles from H, and add description text'
def cat_meta_desc():    
    for x in range(1, len(cats['I'])):
        cats['I'][x].value = str(str(
            cats['H'][x].value) + ' - Langlebige Dentalinstrumente und Zubehör, exklusiv erhältlich bei Ukens Dental')
    print('6.) Column I: Adjusted Meta Descriptions. ')

# Copy column F to B for absolute image paths
def cat_headimg_to_catimg():
    
    for x in range(1, len(cats['A'])):
        cats['B'][x].value = str(cats['F'][x].value)
        # print(str(cats['B'][x].value))
        # print('7.) Column B: now same abs Paths as F')


# save of all changes to new file
def cat_save_file():
    cats_wb.save('files\cats_edited.xlsx')

def runCatFunction():
    cat_meta_title()
    cat_corr_parents()
    cat_friendly_urls()
    cat_img_links()
    cat_corr_cat_ids()
    cat_meta_desc()
    cat_headimg_to_catimg()

    cat_save_file()

runCatFunction()


# End Category Functions

#------------------------------------------------

# Beginning image functions

#wb_imgs = openpyxl.load_workbook('files\products_images.xlsx')
#imgs = wb_imgs['Sheet1']

# def img_rename():